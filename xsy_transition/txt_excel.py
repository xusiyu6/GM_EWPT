import openpyxl
import re

from openpyxl.utils import get_column_letter


def excel_create():
    # 创建一个新的Excel工作表
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.cell(row=1, column=1, value='M1')
    sheet.cell(row=1, column=2, value='M5')
    sheet.cell(row=1, column=3, value='lam2')
    sheet.cell(row=1, column=4, value='lam4')
    # 保存Excel文件
    wb.save('result.xlsx')


def excel_fill():
    # 填写Excel工作表

    wb = openpyxl.load_workbook('result.xlsx')
    sheet = wb.active
    row_num = 1
    column_num = 1

    # 打开txt文件并逐行读取内容
    with open('result.txt', 'r') as file:
        for line in file:
            content = line

            m1_match = re.search(r'M1= (\d+\.\d+)', content)
            if m1_match:
                row_num += 3
                column_num = 1
                sheet.cell(row=row_num, column=column_num, value=m1_match.group(1))
                column_num += 1

            m5_match = re.search(r'M5= (\d+\.\d+)', content)
            if m5_match:
                sheet.cell(row=row_num, column=column_num, value=m5_match.group(1))
                column_num += 1

            lam2_match = re.search(r'lam2= (\d+\.\d+)', content)
            if lam2_match:
                sheet.cell(row=row_num, column=column_num, value=lam2_match.group(1))
                column_num += 1

            lam2ng_match = re.search(r'lam2= -(\d+\.\d+)', content)
            if lam2ng_match:
                sheet.cell(row=row_num, column=column_num, value='-' + lam2ng_match.group(1))
                column_num += 1

            lam4_match = re.search(r'lam4= (\d+\.\d+)', content)
            if lam4_match:
                sheet.cell(row=row_num, column=column_num, value=lam4_match.group(1))
                column_num += 1

            x_match = re.search(r'x = \[(.*?)]', content)
            if x_match:
                column_num = 5
                while sheet.cell(row=row_num, column=column_num).value:
                    row_num += 1
                sheet.cell(row=row_num, column=column_num, value='x=[' + x_match.group(1) + ']')
                column_num += 1

            t_match = re.search(r't = (\d+\.\d+)', content)
            if t_match:
                sheet.cell(row=row_num, column=column_num, value='t=' + t_match.group(1))
                column_num += 1

            up_down_match = re.search(r'up|down', content)
            if up_down_match:
                sheet.cell(row=row_num, column=column_num, value=up_down_match.group(0))
                column_num += 1

            t0_match = re.search(r't0 = (\d+\.\d+)', content)
            if t0_match:
                sheet.cell(row=row_num, column=column_num, value='t0=' + t0_match.group(1))
                column_num += 1
            else:
                t0_nofloat_match = re.search(r't0 = (\d+)', content)
                if t0_nofloat_match:
                    sheet.cell(row=row_num, column=column_num, value='t0=' + t0_nofloat_match.group(1))
                    column_num += 1
    # 保存Excel文件
    wb.save('result.xlsx')


def excel_modify():
    # 修改Excel工作表

    wb = openpyxl.load_workbook('result.xlsx')
    sheet = wb.active
    dims = {}

    # 遍历表格数据，获取自适应列宽数据
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
    for col, value in dims.items():
        # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
        sheet.column_dimensions[get_column_letter(col)].width = value + 2

    # 保存Excel文件
    wb.save('result.xlsx')


excel_create()
excel_fill()
excel_modify()
